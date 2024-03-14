import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import calendar
import os
from openpyxl.drawing.image import Image
import logging

logging.basicConfig(filename='analysis.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def load_datasets():
    try:
        df = pd.read_csv('review_dataset.csv')
        df2 = pd.read_csv('orders_2016-2020_Dataset.csv')
        logging.info("Datasets loaded successfully.")
        return df, df2
    except Exception as e:
        logging.error(f"Error loading datasets: {str(e)}")
        raise

def clean_datasets(df, df2):
    try:
        # Cleaning review dataset
        df['stars'].fillna(0, inplace=True)
        df['status'].fillna('Not Reviewed', inplace=True)
        
        # Cleaning order dataset
        for index, row in df2.iterrows():
            if pd.isna(row['Shipping State']):
                matching_row = df2[(df2['Shipping City'] == row['Shipping City']) & (~df2['Shipping State'].isna())]
                if not matching_row.empty:
                    df2.at[index, 'Shipping State'] = matching_row.iloc[0]['Shipping State']

        df2['Order Date and Time Stamp'] = pd.to_datetime(df2['Order Date and Time Stamp'], format='%d-%m-%Y %H:%M:%S %z')
        df2['Total'] = df2['Total'].str.replace('₹', '').str.replace(',', '')
        df2['Month'] = df2['Order Date and Time Stamp'].dt.month
        df2['Year'] = df2['Order Date and Time Stamp'].dt.year

        logging.info("Datasets cleaned successfully.")
    except Exception as e:
        logging.error(f"Error cleaning datasets: {str(e)}")
        raise

def visualize_star_ratings_distribution(df):
    try:
        dfz = df[df['stars'] != 0]
        review_groups = dfz.groupby('stars').size()
        total_reviews = review_groups.sum()
        percentages = (review_groups / total_reviews) * 100

        ax = review_groups.plot(kind='bar', title='Distribution of Star Ratings')

        for i in range(len(review_groups)):
            plt.text(i, review_groups[i], f"{percentages[i]:.2f}%", ha='center', va='bottom')

        plt.savefig('plot_image.png', dpi=300)
        review_groups.to_excel('distribution_of_star_ratings.xlsx', header=['Count'])

        wb = openpyxl.load_workbook('distribution_of_star_ratings.xlsx')
        ws = wb.active
        img = Image('plot_image.png')
        ws.add_image(img, 'A10')
        wb.save('distribution_of_star_ratings.xlsx')
        plt.show()
        logging.info("Star ratings distribution visualized and saved successfully.")
    except Exception as e:
        logging.error(f"Error visualizing star ratings distribution: {str(e)}")
        raise

def visualize_payment_distribution(df2, excel_filename):
    # Data segregation based on payment method
    df2['Payment Method'] = df2['Payment Method'].str.replace(r'[0-9.,]+', '', regex=True)
    payment_method_groups = df2.groupby('Payment Method').size()

    # Data visualization
    payment_method_groups.plot(kind='bar', title='Distribution of Payment Methods')
    plt.savefig('plot_image.png', dpi=300)

    # Check if Excel file already exists, if yes, delete it
    if os.path.exists(excel_filename):
        os.remove(excel_filename)

    # Save the data to an Excel file
    payment_method_groups.to_excel(excel_filename)

    # Load the Excel file and add the image
    wb = openpyxl.Workbook()
    ws = wb.active
    img = Image('plot_image.png')
    ws.add_image(img, 'A10')

    # Save the Excel file with the image
    wb.save(excel_filename)

    # Show the plot
    plt.show()

def generate_state_analysis_plots(df2):
    try:
        # Plot top consumer states
        state_counts = df2['Shipping State'].value_counts()
        top_states = state_counts.head(10)
        top_states.plot(kind='bar', title='Top Consumer States')
        plt.xlabel('State')
        plt.ylabel('Number of Consumers')
        plt.xticks(rotation=45)
        plt.savefig('top_consumer_states_plot.png', dpi=300)
        plt.close()

        # Calculate state revenue and consumers
        state_revenue = df2.groupby('Shipping State')['Total'].sum()
        state_consumers = df2['Shipping State'].value_counts()
        total_consumers = state_consumers.sum()
        state_percentages = (state_consumers / total_consumers) * 100
        state_data = pd.DataFrame({'Total Revenue': state_revenue, 
                                'Number of Consumers': state_consumers, 
                                'Percentage of Consumers': state_percentages})

        # Plot total revenue by state
        plt.figure(figsize=(10, 6))
        plt.bar(state_data.index, state_data['Total Revenue'], color='blue', alpha=0.7)
        plt.xlabel('State')
        plt.ylabel('Total Revenue')
        plt.title('Total Revenue by State')
        plt.xticks(rotation=45)
        formatter = ticker.StrMethodFormatter('{x:,.0f}')
        plt.gca().yaxis.set_major_formatter(formatter)
        plt.savefig('total_revenue_plot.png', dpi=300)
        plt.close()

        # Plot number of consumers by state
        plt.figure(figsize=(10, 6))
        plt.bar(state_data.index, state_data['Number of Consumers'], color='green', alpha=0.7)
        plt.xlabel('State')
        plt.ylabel('Number of Consumers')
        plt.title('Number of Consumers by State')
        plt.xticks(rotation=45)
        plt.savefig('num_consumers_plot.png', dpi=300)
        plt.close()

        # Plot percentage of consumers by state
        plt.figure(figsize=(10, 6))
        plt.bar(state_data.index, state_data['Percentage of Consumers'], color='orange', alpha=0.7)
        plt.xlabel('State')
        plt.ylabel('Percentage of Consumers')
        plt.title('Percentage of Consumers by State')
        plt.xticks(rotation=45)
        plt.savefig('percentage_consumers_plot.png', dpi=300)
        plt.close()

        # Save data to Excel file
        if os.path.exists('top_consumer_state_data.xlsx'):
            os.remove('top_consumer_state_data.xlsx')
        state_data.to_excel('top_consumer_state_data.xlsx')

        # Load Excel file and add images
        wb = openpyxl.load_workbook('top_consumer_state_data.xlsx')
        ws = wb.active

        # Add total revenue plot image
        total_revenue_img = Image('total_revenue_plot.png')
        ws.add_image(total_revenue_img, 'A10')

        # Add number of consumers plot image
        num_consumers_img = Image('num_consumers_plot.png')
        ws.add_image(num_consumers_img, 'A150')

        # Add percentage of consumers plot image
        percentage_consumers_img = Image('percentage_consumers_plot.png')
        ws.add_image(percentage_consumers_img, 'A290')

        # Save Excel file with images
        wb.save('top_consumer_state_data.xlsx')

    except Exception as e:
        print(f"An error occurred: {str(e)}")

def generate_city_analysis_plots(df2):
    try:
        # Group by 'Shipping City' and calculate total revenue
        city_revenue = df2.groupby('Shipping City')['Total'].sum()

        # Group by 'Shipping City' and count the number of occurrences of each city
        city_consumers = df2['Shipping City'].value_counts()

        # Calculate the percentage of consumers in each city
        total_consumers = city_consumers.sum()
        city_percentages = (city_consumers / total_consumers) * 100

        # Create a DataFrame to store the results
        city_data = pd.DataFrame({'Total Revenue': city_revenue, 
                                  'Number of Consumers': city_consumers, 
                                  'Percentage of Consumers': city_percentages})

        # Select the top 10 cities by revenue
        top_10_revenue = city_data.nlargest(10, 'Total Revenue')

        # Plot the Total Revenue for the top 10 cities
        plt.figure(figsize=(10, 6))
        plt.bar(top_10_revenue.index, top_10_revenue['Total Revenue'], color='blue', alpha=0.7)
        plt.xlabel('City')
        plt.ylabel('Total Revenue')
        plt.title('Total Revenue for Top 10 Cities by Revenue')
        plt.xticks(rotation=45)
        formatter = ticker.StrMethodFormatter('{x:,.0f}')  # Format as integer with thousand separators
        plt.gca().yaxis.set_major_formatter(formatter)
        plt.savefig('total_revenue_plot.png', dpi=300)  # Save the plot image
        plt.close()

        # Plot the Number of Consumers for the top 10 cities
        top_10_consumers = city_data.nlargest(10, 'Number of Consumers')
        plt.figure(figsize=(10, 6))
        plt.bar(top_10_consumers.index, top_10_consumers['Number of Consumers'], color='green', alpha=0.7)
        plt.xlabel('City')
        plt.ylabel('Number of Consumers')
        plt.title('Number of Consumers for Top 10 Consumer Cities')
        plt.xticks(rotation=45)
        plt.savefig('num_consumers_plot.png', dpi=300)
        plt.close()

        # Save the data to an Excel file
        if os.path.exists('Top_consumer_city.xlsx'):
            os.remove('Top_consumer_city.xlsx')
        city_data.to_excel('Top_consumer_city.xlsx')

        # Load the Excel file and add the images
        wb = openpyxl.load_workbook('Top_consumer_city.xlsx')
        ws = wb.active

        # Add the Total Revenue plot image
        total_revenue_img = Image('total_revenue_plot.png')
        ws.add_image(total_revenue_img, 'A10')

        # Add the Number of Consumers plot image
        num_consumers_img = Image('num_consumers_plot.png')
        ws.add_image(num_consumers_img, 'A150')

        # Save the Excel file with the images
        wb.save('Top_consumer_city.xlsx')

    except Exception as e:
        print(f"An error occurred: {str(e)}")

def plot_top_selling_categories(df, df2):
    try:
        merged_df = pd.merge(df, df2, left_on='product_name', right_on='LineItem Name')
        
        # Group by the product categories and sum the quantity sold and revenue for each category
        category_data_qty = merged_df.groupby('category')['LineItem Qty'].sum()
        category_data_revenue = merged_df.groupby('category')['Total'].sum()
        
        # Combine the aggregated data into a single DataFrame
        category_data = pd.DataFrame({'Quantity Sold': category_data_qty, 'Total Revenue': category_data_revenue})
        
        # Plot the top selling product categories by quantity sold
        top_categories_qty = category_data['Quantity Sold'].nlargest(10)
        top_categories_qty.plot(kind='bar', figsize=(10, 6), color='blue')
        plt.xlabel('Product Category')
        plt.ylabel('Quantity Sold')
        plt.title('Top Selling Product Categories by Quantity Sold')
        plt.xticks(rotation=45)
        plt.savefig('Top_Selling_Product_Categories.png', dpi=300)
        plt.show()

        # Plot the top selling product categories by total revenue
        top_categories_revenue = category_data['Total Revenue'].nlargest(10)
        top_categories_revenue.plot(kind='bar', figsize=(10, 6), color='green')
        plt.xlabel('Product Category')
        plt.ylabel('Total Revenue')
        plt.title('Top Selling Product Categories by Total Revenue')
        plt.xticks(rotation=45)
        plt.savefig('Top_selling_product_totalrevenue.png', dpi=300)
        plt.show()

        if os.path.exists('Top_product.xlsx'):
            os.remove('Top_product.xlsx')
        category_data.to_excel('Top_product.xlsx', sheet_name='Top_Product_Categories', index=True)

        # Load the Excel file and add the images
        wb = openpyxl.load_workbook('Top_product.xlsx')
        ws = wb.active

        # Add the Quantity Sold plot image
        qty_img = Image('Top_Selling_Product_Categories.png')
        ws.add_image(qty_img, 'A10')

        # Add the Total Revenue plot image
        revenue_img = Image('Top_selling_product_totalrevenue.png')
        ws.add_image(revenue_img, 'F10')

        # Save the Excel file with the images
        wb.save('Top_product.xlsx')
    except Exception as e:
        print(f"An error occurred: {str(e)}")


def plot_category_ratings(df):
    try:
        category_review_counts = df.groupby(['category', 'stars']).size().unstack(fill_value=0)
        category_review_counts['total_count'] = category_review_counts.sum(axis=1)
        top_categories = category_review_counts.nlargest(13, 'total_count')
        df_top_categories = df[df['category'].isin(top_categories.index)]

        plt.figure(figsize=(12, 8))
        ax = df_top_categories.groupby(['category', 'stars']).size().unstack().plot(kind='bar', stacked=True)
        ax.set_xlabel('Product Category', fontsize=12)
        ax.set_ylabel('Count', fontsize=12)
        plt.title('Distribution of Ratings for Top 13 Product Categories')
        plt.xticks(rotation=45, fontsize=10)
        plt.yticks(fontsize=10)
        plt.legend(title='Rating', bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.tight_layout()

        plt.savefig('Rating_of_Category.png', dpi=300)

        category_review_counts.to_excel('Top_category.xlsx')

        wb = openpyxl.load_workbook('Top_category.xlsx')
        ws = wb.active

        total_revenue_img = Image('Rating_of_Category.png')
        ws.add_image(total_revenue_img, 'A10')

        wb.save('Top_category.xlsx')
        plt.show()

    except Exception as e:
        print(f"An error occurred: {str(e)}")   

def plot_orders_and_revenue_per_month(df2):
    try:
        df2['Order Date and Time Stamp'] = pd.to_datetime(df2['Order Date and Time Stamp'])
        df2['Year'] = df2['Order Date and Time Stamp'].dt.year

        orders_per_month_per_year = df2.groupby(['Year', df2['Order Date and Time Stamp'].dt.month])['Order Date and Time Stamp'].count()
        revenue_per_month_per_year = df2.groupby(['Year', df2['Order Date and Time Stamp'].dt.month])['Total'].sum()

        for year in df2['Year'].unique():
            plt.figure(figsize=(10, 6))
            
            orders_data = orders_per_month_per_year.loc[year]
            revenue_data = revenue_per_month_per_year.loc[year]

            plt.subplot(2, 1, 1)
            orders_data.plot(kind='bar', color='blue', alpha=0.7)
            plt.xlabel('Month')
            plt.ylabel('Number of Orders')
            plt.title(f'Number of Orders Per Month in {year}')
            plt.xticks(range(0, 12), [calendar.month_abbr[i] for i in range(1, 13)], rotation=45)
            plt.tight_layout()

            plt.subplot(2, 1, 2)
            revenue_data.plot(kind='bar', color='green', alpha=0.7)
            plt.xlabel('Month')
            plt.ylabel('Revenue')
            plt.title(f'Revenue Per Month in {year}')
            plt.xticks(range(0, 12), [calendar.month_abbr[i] for i in range(1, 13)], rotation=45)
            plt.tight_layout()

            plt.savefig(f'Revenue_and_Orders_{year}.png', dpi=300)
            plt.close()

            # Save data to Excel file
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'Month'
            ws['B1'] = 'Number of Orders'
            ws['C1'] = 'Revenue'

            for i, (month, orders, revenue) in enumerate(zip(orders_data.index, orders_data.values, revenue_data.values), start=2):
                ws[f'A{i}'] = calendar.month_abbr[month]
                ws[f'B{i}'] = orders
                ws[f'C{i}'] = revenue

            wb.save(f'Revenue_and_Orders_{year}.xlsx')

    except Exception as e:
        print(f"An error occurred: {str(e)}") 



def plot_orders_and_reviews_per_month_and_year(df, df2):
    try:
        merged_df = pd.merge(df, df2, left_on='product_name', right_on='LineItem Name')
        category_data = merged_df.groupby('category').agg({'LineItem Qty': 'sum', 'Total': 'sum'})
        merged_df['stars_numeric'] = merged_df['stars'].str.extract(r'(\d+\.?\d*)').astype(float)

        order_review_data = merged_df.groupby(['Year', 'Month']).agg({'Order #': 'count', 'stars_numeric': 'mean'}).reset_index()

        plt.figure(figsize=(12, 6))
        plt.plot(order_review_data['Month'], order_review_data['Order #'], marker='o', label='Number of Orders')
        plt.plot(order_review_data['Month'], order_review_data['stars_numeric'], marker='s', label='Average Review')
        plt.xlabel('Month')
        plt.ylabel('Count / Review')
        plt.title('Number of Orders and Average Review Per Month Per Year')
        plt.legend()
        plt.xticks(range(1, 13), [calendar.month_abbr[i] for i in range(1, 13)])  # Convert month numbers to names
        plt.grid(True)
        plt.savefig('Orders_and_Reviews_Per_Month.png', dpi=300)
        plt.close()

        yearly_order_review_data = merged_df.groupby(['Year']).agg({'Order #': 'count', 'stars_numeric': 'mean'}).reset_index()

        plt.figure(figsize=(10, 6))
        plt.plot(yearly_order_review_data['Year'], yearly_order_review_data['Order #'], marker='o', label='Number of Orders')
        plt.plot(yearly_order_review_data['Year'], yearly_order_review_data['stars_numeric'], marker='s', label='Average Review')
        plt.xlabel('Year')
        plt.ylabel('Count / Review')
        plt.title('Number of Orders and Average Review Per Year')
        plt.legend()
        plt.xticks(yearly_order_review_data['Year'])  # Set the x-axis ticks to the years
        plt.grid(True)
        plt.savefig('Orders_and_Reviews_Per_Year.png', dpi=300)
        plt.close()

        # Save data to Excel file
        with pd.ExcelWriter('Orders_and_Reviews.xlsx', engine='openpyxl') as writer:
            order_review_data.to_excel(writer, sheet_name='Orders_Reviews_Monthly', index=False)
            yearly_order_review_data.to_excel(writer, sheet_name='Orders_Reviews_Yearly', index=False)

    except Exception as e:
        print(f"An error occurred: {str(e)}")

def plot_orders_by_year_day_part(df2):
    try:
        years = df2['Year'].unique()
        df2['Hour'] = df2['Order Date and Time Stamp'].dt.hour

        # Define a function to categorize the hour into parts of a day
        def categorize_hour(hour):
            if 0 <= hour < 6:
                return 'Night'
            elif 6 <= hour < 12:
                return 'Morning'
            elif 12 <= hour < 18:
                return 'Afternoon'
            else:
                return 'Evening'

        # Apply the function to categorize the hour
        df2['Day Part'] = df2['Hour'].apply(categorize_hour)

        # Group by the parts of the day and count the number of orders
        orders_by_day_part = df2.groupby('Day Part')['Order #'].count()

        with pd.ExcelWriter('Orders_By_Year_Day_Part.xlsx', engine='openpyxl') as writer:
            for year in years:
                plt.figure(figsize=(8, 5))
                orders_by_day_part.plot(kind='bar', color='blue', alpha=0.7)
                plt.xlabel('Part of Day')
                plt.ylabel('Number of Orders')
                plt.title(f'Number of Orders Across Parts of a Day for {year}')
                plt.xticks(rotation=45)
                plt.grid(axis='y')
                plt.tight_layout()

                # Save plot to Excel
                plt.savefig(f'Orders_By_Day_Part_{year}.png', dpi=300)
                plt.close()

                # Save data to Excel
                orders_by_day_part.to_excel(writer, sheet_name=f'Orders_{year}')

        print("Data and graphs saved successfully!")

    except Exception as e:
        print(f"An error occurred: {str(e)}")


def run_all_analysis():
    try:
        df, df2 = load_datasets()
        clean_datasets(df, df2)
        visualize_star_ratings_distribution(df)
        # Call other analysis functions here
        logging.info("All analysis functions executed successfully.")
    except Exception as e:
        logging.exception(f"An error occurred: {str(e)}")

def main():
    try:
        df, df2 = load_datasets()
        clean_datasets(df, df2)

        while True:
            print("Press 1 to visualize star ratings distribution.")
            print("Press 2 to visualize payment distribution.")
            print("Press 3 to analyze top consumer states.")
            print("Press 4 to analyze top consumer cities.")
            print("Press 5 to plot top selling categories.")
            print("Press 6 to plot category ratings.")
            print("Press 7 to plot orders and revenue per month.")
            print("Press 8 to plot orders and reviews per month and year.")
            print("Press 9 to plot orders by year and day part.")
            print("Press 10 to run all analysis functions.")
            print("Press 11 to exit.")
            choice = input("Enter your choice: ")

            if choice == '1':
                visualize_star_ratings_distribution(df)
                break
            elif choice == '2':
                excel_filename = 'distribution_of_payment_method.xlsx'
                visualize_payment_distribution(df2, excel_filename)
                break
            elif choice == '3':
                generate_state_analysis_plots(df2)
                break
            
            elif choice == '4':
                generate_city_analysis_plots(df2)
                break
            elif choice == '5':
                plot_top_selling_categories(df,df2)
                break
            elif choice == '6':
                plot_category_ratings(df)
                break
            elif choice == '7':
                plot_orders_and_revenue_per_month(df2)
                break
            elif choice == '8':
                plot_orders_and_reviews_per_month_and_year(df, df2)
                break
            elif choice == '9':
                plot_orders_by_year_day_part(df2)
                break
            elif choice == '10':
                run_all_analysis(df, df2)
                break
            elif choice == '11':
                        break
            else:
                print("Invalid choice. Please try again.")
                    
                    
                   
    except Exception as e:
        logging.exception(f"An unexpected error occurred: {str(e)}")

def run_all_analysis(df, df2):
    visualize_star_ratings_distribution(df)
    excel_filename = 'distribution_of_payment_method.xlsx'
    visualize_payment_distribution(df2, excel_filename)
    generate_state_analysis_plots(df2)
    # Add calls to other analysis functions

if __name__ == "__main__":
    main()
